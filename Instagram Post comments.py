from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
import time
import pandas as pd
from openpyxl import load_workbook
import pyautogui
import win10toast
import smtplib
import gtts
import pyttsx3
import os
import re
from selenium.webdriver.chrome.options import Options

#########################################################################################################################################################################################
#                   # Mentioning chromedriver that it is headless
###########################################################################################################################################################################################

options = Options()
options.add_argument("--headless")
options.add_argument('window-size=1920x1080')
options.add_argument("--start-maximized")

#########################################################################################################################################################################################
#                    # Providing the chrome driver executable path
###########################################################################################################################################################################################

driver = webdriver.Chrome('E:\Program Files\chromedriver_win32\chromedriver.exe')
time.sleep(10)
driver.get("https://www.instagram.com/")
driver.maximize_window()

filename="Comments.xlsx"
workbook=load_workbook(filename="Comments.xlsx")
sheet=workbook.active
row_number=sheet.max_row

Done='Hey Mani Activity iiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiiis doneeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee'
Error='Hey Mani Errorrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrr ocurred'
welcome_message='Hey Mani Good Morning mera dost!!!!!!!!!!!!'
wait=WebDriverWait(driver,2)



def login_details():
    workbook1= pd.read_excel(r'E:\Manikanta\Instagram\Credintials.xlsx')
    uname=workbook1.Uname[1]
    pwd= workbook1.PWD[1]
    return uname,pwd

def voice(message):
   engine = pyttsx3.init()
   engine.say(message)
   engine.runAndWait()

def login():
    time.sleep(5)
    uname,pwd = login_details()
# username enter
    ubutton = driver.find_element_by_xpath("//input[@name='username']")
    time.sleep(1)
    ubutton.send_keys(uname)

    time.sleep(1)
#password enter
    pbutton = driver.find_element_by_xpath("//input[@name='password']")
    time.sleep(1)
    pbutton.send_keys(pwd)

#login button click

    login_button = driver.find_element_by_xpath("//Button[@type='submit']")
        #"/html/body/div[1]/section/main/article/div[2]/div[1]/div/form/div[4]/button/div")
    login_button.click()
    time.sleep(7)
    print('login completed')
    mouse_move()

def instagram_desktop_notification():
    #########################################################################################################################################################################################
    #                    # Notification for desktop notification
    ###########################################################################################################################################################################################


    notification1 = driver.find_element_by_xpath("/html/body/div[1]/section/main/div/div/div/div/button")
    time.sleep(2)
    notification1.click()
    time.sleep(5)

# Notification for desktop notification

    notification = driver.find_element_by_xpath("/html/body/div[4]/div/div/div/div[3]/button[2]")
    time.sleep(2)
    notification.click()
    time.sleep(5)
    print('notifications removed')
    mouse_move()

def mouse_move():
    #########################################################################################################################################################################################
    #                    # For Mouse movement
    ###########################################################################################################################################################################################

    #pyautogui.moveTo(100, 100, duration=1)
    #pyautogui.moveTo(10, 10, duration=1)
    pyautogui.press('volumedown')
    time.sleep(1)
    pyautogui.press('volumeup')
    time.sleep(5)

def updating_comments_in_excel(id,comments,row_number):
    count=len(id)
    print(count)
    i=row_number
    for k in id:
        i = 1 + int(i)
        try:
            if i != count:
                column_name1='A' +str(i)
                cell1=sheet[column_name1]
                cell1.value=id[int(i)]
                column_name = 'B' + str(i)
                cell = sheet[column_name]
                cell.value = comments[int(i)]
                workbook.save(filename = filename)
            else:
                break
        except:
                 continue
                 workbook.save(filename = filename)


    workbook.save(filename = filename)


def getting_post_links():
    #########################################################################################################################################################################################
    #                    # getting links from the profile
    ###########################################################################################################################################################################################

        link = 'https://www.instagram.com/amaramanikanta_photography/'
        driver.get(link)
        ylinks=[]

        for i in range(2):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            a_tag_links = driver.find_elements_by_tag_name('a')
            # count=print('total no of posts from'+ g+' :'+len(a_tag_links))
            post_links = [m.get_attribute('href') for m in a_tag_links if '.com/p/' in m.get_attribute('href')]
            for i in post_links:
                if i not in ylinks:
                    ylinks.append(i)

        print("no of posts",len(ylinks))
        return ylinks

def getting_comments(mani):
    #########################################################################################################################################################################################
    #                    # Getting comments from the posts
    ###########################################################################################################################################################################################

    links=mani
    id = []
    comments = []
    for j in links:
        driver.get(j)
       # driver.maximize_window()

        for i in range(10):
           try:
            load_more_comments_button = driver.find_element_by_xpath("//*[contains(@class,'dCJp8 afkep')]")
            driver.execute_script("arguments[0].scrollIntoView();",load_more_comments_button)
            if wait.until(ec.visibility_of_element_located((By.XPATH,"//*[contains(@class,'dCJp8 afkep')]"))) :
                driver.execute_script("window.scrollTo(0, 0);")
                #time.sleep(1)
                driver.find_element_by_xpath("//*[contains(@class,'dCJp8 afkep')]").click()
                time.sleep(2)
           except:
               continue
        comments_xpaths = driver.find_elements_by_xpath("//*[contains(@class,'C4VMK')]")
        print(len(comments_xpaths))

        for i in comments_xpaths:
            string=i.text
            remove_text='1d1 likeReply'
            if (string.find('amaramanikanta_photography') == -1):
                res=re.sub(remove_text,"",string)
                final=(res.rstrip().split('\n'))
                id.append(final[0])
                comments.append(final[1])
    row_number = sheet.max_row
    print(len(id))
    updating_comments_in_excel(id,comments,row_number)









login()
time.sleep(1)

instagram_desktop_notification()
time.sleep(1)

post_links = getting_post_links()
time.sleep(1)
getting_comments(post_links)
